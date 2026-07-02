import logging
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse(file_path: str) -> dict:
    try:
        if file_path.endswith(".xls"):
            return _parse_xls(file_path)
        return _parse_xlsx(file_path)
    except Exception as exc:
        logger.error("Excel parse error: %s", exc)
        return {"error": str(exc)}


def _parse_xlsx(file_path: str) -> dict:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[name] = {"issue": "Empty sheet"}
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        sample = [dict(zip(headers, r)) for r in rows[1:6]]
        sheets[name] = {
            "row_count": ws.max_row - 1,
            "column_count": ws.max_column,
            "columns": headers,
            "sample": sample,
        }
    wb.close()
    return {"file_type": "Excel (xlsx/xlsm)", "sheets": sheets}


def _parse_xls(file_path: str) -> dict:
    df = pd.read_excel(file_path, engine="xlrd", nrows=5)
    full = pd.read_excel(file_path, engine="xlrd")
    return {
        "file_type": "Excel (xls)",
        "shape": {"rows": full.shape[0], "columns": full.shape[1]},
        "schema": {col: str(dtype) for col, dtype in full.dtypes.items()},
        "sample": df.to_dict(orient="records"),
        "issues": {
            col: int(full[col].isna().sum())
            for col in full.columns
            if full[col].isna().any()
        },
    }
